#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side


def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Generate a formatted Excel file of supplementary tables."
    )
    parser.add_argument(
        "-s",
        "--supp-table-list",
        default="conf/suppTableList.txt",
        help="Path to list of supplementary tables (default: conf/suppTableList.txt)"
    )
    parser.add_argument(
        "-c",
        "--contents",
        default="conf/contents.tsv",
        help="Path to contents file (default: conf/contents.tsv)"
    )
    parser.add_argument(
        "-g",
        "--glossary",
        default="conf/glossary.tsv",
        help="Path to glossary file (default: conf/glossary.tsv)"
    )
    parser.add_argument(
        "-o",
        "--output",
        default="results/suppTables.xlsx",
        help="Path to output Excel file (default: results/suppTables.xlsx)"
    )
    return parser.parse_args()


def main():
    args = parse_args()

    # Read the list of tables; now assumes each cell already contains the full path
    tables = pd.read_csv(args.supp_table_list, header=None)[0]

    # Prepend Contents and Glossary to the list
    table_paths = pd.concat([pd.Series([args.contents, args.glossary]), tables])

    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        for j, table_path in enumerate(table_paths):

            df = pd.read_csv(table_path, sep="\t", index_col=0)

            # Sheet naming logic
            if j == 0:
                sheet_name = "Contents"
            elif j == 1:
                sheet_name = "Glossary"
            else:
                sheet_name = os.path.splitext(os.path.basename(table_path))[0]

            df.to_excel(writer, sheet_name=sheet_name)

    # Formatting section
    thin_border = Border(bottom=Side(style="thin"))
    workbook = load_workbook(args.output)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # clear formatting
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(bold=False)
                cell.border = Border()

        # apply header formatting
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # auto column width
        for col in sheet.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0) + 2
            col_idx = col[0].column if col[0].column is not None else 1
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = max(max_length, 12)

    workbook.save(args.output)


if __name__ == "__main__":
    main()

