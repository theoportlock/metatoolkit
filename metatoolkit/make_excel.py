#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side


def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Generate a formatted Excel workbook from multiple TSV tables."
    )

    parser.add_argument(
        "-l",
        "--table-list",
        required=True,
        nargs="+",
        help=(
            "Table inputs: list file(s), TSV file(s), and/or wildcards "
            "(e.g. tables/*.tsv)"
        )
    )

    parser.add_argument(
        "--toc",
        help="Optional table-of-contents TSV file"
    )

    parser.add_argument(
        "--glossary",
        help="Optional glossary TSV file"
    )

    parser.add_argument(
        "-o",
        "--output",
        default="results/output.xlsx",
        help="Path to output Excel file (default: results/output.xlsx)"
    )

    return parser.parse_args()


def resolve_table_paths(inputs):
    """
    Resolve table paths from:
      - list files (one path per line)
      - direct TSV files
      - wildcards
    """
    paths = []

    for item in inputs:
        expanded = glob.glob(item)

        if expanded:
            for path in expanded:
                if os.path.isfile(path) and path.endswith(".tsv"):
                    paths.append(path)
                elif os.path.isfile(path):
                    # assume list file
                    listed = pd.read_csv(path, header=None)[0].tolist()
                    paths.extend(listed)
        else:
            raise FileNotFoundError(f"No files match input: {item}")

    # Remove duplicates while preserving order
    paths = list(dict.fromkeys(paths))

    if not paths:
        raise ValueError("No valid table files found.")

    return paths


def write_tables(writer, table_paths, toc=None, glossary=None):
    """Write all tables to an Excel writer."""
    ordered_tables = []

    if toc:
        ordered_tables.append(("Contents", toc))

    if glossary:
        ordered_tables.append(("Glossary", glossary))

    for path in table_paths:
        sheet_name = os.path.splitext(os.path.basename(path))[0][:31]
        ordered_tables.append((sheet_name, path))

    for sheet_name, path in ordered_tables:
        df = pd.read_csv(path, sep="\t", index_col=0)
        df.to_excel(writer, sheet_name=sheet_name)


def format_workbook(path):
    """Apply consistent formatting to all sheets."""
    thin_border = Border(bottom=Side(style="thin"))
    workbook = load_workbook(path)

    for sheet in workbook.worksheets:

        # Clear formatting
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(bold=False)
                cell.border = Border()
                cell.alignment = Alignment(horizontal="general")

        # Header formatting
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Auto column widths
        for col in sheet.columns:
            max_length = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=0
            ) + 2

            col_idx = col[0].column
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = max(max_length, 12)

    workbook.save(path)


def main():
    args = parse_args()

    table_paths = resolve_table_paths(args.table_list)

    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        write_tables(
            writer,
            table_paths=table_paths,
            toc=args.toc,
            glossary=args.glossary
        )

    format_workbook(args.output)


if __name__ == "__main__":
    main()

